import csv
import random
from PIL import Image, ImageOps
from PRNU_Extraction import noise_extract as prnu
from PRNU_Extraction import check_orientation, cut_ctr
from Cross_Correlation_Functions import pce, crosscorr_2d
import numpy as np
import openpyxl as pyxl
import time
from datetime import datetime


labels = {} # Path : Device
paths = []
devices = {'D3500': 0, 'Drone': 1, 'iPhone 8' : 2, 'iPhone13' : 3, 'Samsung Galaxy S7 A' : 4, 'Samsung Galaxy S7 B' : 5, 0 : 'D3500', 1 : 'Drone', 2 : 'iPhone 8', 3 : 'iPhone 13', 4 : 'Samsung Galaxy S7 A', 5 : 'Samsung Galaxy S7 B' }
full_dataset = [['Image Path', 'Actual Class', 'Predicted Class', 'D3500', 'Drone', 'iPhone 8', 'iPhone 13', 'Galaxy S7 A', 'Galaxy S7 B', 'Result', 'Mode']]

def load_data():
    paths.clear()
    labels.clear()
    with open(r"C:\Users\Elliot\Documents\Comp Sci Year 3\Final Year Project\Classified_Data.csv", mode="r") as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        for row in csv_reader:
            labels[row[0]] = row[1]
            paths.append(row[0])

def load_equal_data(dataset_size):
    paths.clear()
    labels.clear()
    devices_count = {'D3500': 0, 'Drone': 0, 'iPhone 8' : 0, 'iPhone13' : 0, 'Samsung Galaxy S7 A' : 0, 'Samsung Galaxy S7 B' : 0}
    with open(r"C:\Users\Elliot\Documents\Comp Sci Year 3\Final Year Project\Classified_Data.csv", mode="r") as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        for row in csv_reader:
            if devices_count[row[1]] < dataset_size:
                labels[row[0]] = row[1]
                paths.append(row[0])
                devices_count[row[1]] +=1

def create_fold(folds):
    folded_data = []
    random.shuffle(paths)
    fold_size = (len(labels))//folds # Ammount of entries in each fold
    remainder = (len(labels))%folds # Ammount of left of entries
    fold_quant = (len(labels))//fold_size
    counter = 0
    for x in range(fold_quant):
        current_fold = []
        for i in range(fold_size):
            current_fold.append(paths[counter])
            counter +=1
        folded_data.append(current_fold)
    if remainder > 0:
        current_fold = []
        for i in range(remainder):
            current_fold.append(paths[counter])
            counter +=1
        folded_data.append(current_fold)
    return folded_data

def create_fixed_size_fold(fold_size):
    folded_data = []
    random.shuffle(paths)
    fold_count = len(labels)//fold_size
    print(f"{fold_count} Folds Will Be Created, Each With {fold_size} Entries.")
    print("-----------------------------------------------------------------------------------------------------------")
    if fold_count > 0:
        counter = 0
        for i in range(fold_count):
            current_fold = []
            for i in range(fold_size):
                current_fold.append(paths[counter])
                counter += 1
            folded_data.append(current_fold)
    return folded_data

def create_equal_sized_folds(fold_size = 50, device_data_quant = 50):
    folded_data = []
    random.shuffle(paths)
    load_equal_data(device_data_quant) # 50 images of each devices
    # this will create a fold total size of 25*6= 150 images per fold
    this_paths = paths.copy()
    random.shuffle(this_paths)
    fold_quant = device_data_quant//fold_size # 2
    
    print(f"{fold_quant} Folds Will Be Created, Each With {fold_size*6} Entries.")
    print("-----------------------------------------------------------------------------------------------------------")

    if fold_quant > 0:
        for i in range(fold_quant):
            counter = 0
            devices_count = {'D3500': 0, 'Drone': 0, 'iPhone 8' : 0, 'iPhone13' : 0, 'Samsung Galaxy S7 A' : 0, 'Samsung Galaxy S7 B' : 0}
            current_fold = []
            for image in this_paths:
                if (sum(devices_count.values())) < 300:
                    if devices_count[labels[this_paths[counter]]] < fold_size:
                        current_fold.append(this_paths[counter])
                        devices_count[labels[this_paths[counter]]] += 1
                        counter += 1
                    else:
                        counter+=1
                else:
                    break
            for i in current_fold[:]:
                if i in this_paths:
                    this_paths.remove(i)
            folded_data.append(current_fold)
    return folded_data

def create_no_match_fold(fold_size = 50, device_data_quant = 50):
    device_names = ['D3500', 'Drone', 'iPhone 8', 'iPhone13', 'Samsung Galaxy S7 A', 'Samsung Galaxy S7 B']
    folded_data = []
    random.shuffle(paths)
    load_equal_data(device_data_quant)

    this_paths = paths.copy()
    random.shuffle(this_paths)
    fold_quant = device_data_quant//fold_size 
    

    print(f"{fold_quant} Folds Will Be Created, Each With {fold_size*5} Entries.")
    print("-----------------------------------------------------------------------------------------------------------")

    excluded_device = random.choice(device_names)
    print(f"Excluded Device: {excluded_device}")

    if fold_quant > 0:
        for i in range(fold_quant):
            counter = 0
            devices_count = {'D3500': 0, 'Drone': 0, 'iPhone 8' : 0, 'iPhone13' : 0, 'Samsung Galaxy S7 A' : 0, 'Samsung Galaxy S7 B' : 0}
            current_fold = []
            for image in this_paths:
                if (sum(devices_count.values())) < 251:
                    if labels[this_paths[counter]] == excluded_device and devices_count[labels[this_paths[counter]]] < 1:
                        current_fold.append(this_paths[counter])
                        devices_count[labels[this_paths[counter]]] += 1
                        counter += 1
                    elif labels[this_paths[counter]] == excluded_device:
                        counter +=1
                    elif devices_count[labels[this_paths[counter]]] < fold_size:
                        current_fold.append(this_paths[counter])
                        devices_count[labels[this_paths[counter]]] += 1
                        counter += 1
                    else:
                        counter+=1
                else:
                    break
            for i in current_fold[:]:
                if i in this_paths:
                    this_paths.remove(i)
            folded_data.append(current_fold)
    return folded_data, excluded_device

def get_fold_stats(folds):
     for fold in folds:
        count = [0,0,0,0,0,0]
        for image in fold:
            device = labels[image]
            count[devices[device]] += 1
        print(f"| D3500: {count[0]} | Drone: {count[1]} | iPhone 8: {count[2]} | iPhone 13: {count[3]} | Samsung Galaxy S7 A: {count[4]} | Samsung Galaxy S7 B: {count[5]} |") 
        
def choose_suspect_image(folds):
    # Do things
    new_fold = []
    for fold in folds:
        current_fold = []
        current_fold.append(fold[0])
        del fold[0]
        current_fold.append(fold)
        new_fold.append(current_fold)
    return new_fold

def choose_no_match_suspect(folds, excluded):
    new_fold = []
    for fold in folds:
        current_fold = []
        for i in fold:
            if labels[i] == excluded:
                current_fold.append(i)
                fold.remove(i)
        current_fold.append(fold)
        new_fold.append(current_fold)
    return new_fold

def open_control(paths):
    control_images = []

    for i in paths:
        hold = Image.open(i)
        hold = ImageOps.exif_transpose(hold)
        hold = hold.convert('L')
        control_images.append(check_orientation(hold))
    return control_images

def open_suspect(path):
    hold = Image.open(path)
    hold = ImageOps.exif_transpose(hold)
    hold = hold.convert('L')
    suspect_image = check_orientation(hold)
    suspect_image = np.asarray(suspect_image)
    suspect_image = cut_ctr(suspect_image, (suspect_image.shape))
    return suspect_image

def average(list): 
    if len(list) > 0:
        return sum(list) / len(list)
    else:
        return 0
    
def check_upper_bound(): 
    load_data()
    fold = [paths[1], paths[0]]
    print(paths[1])
    suspect_image_path = fold[0]
    control_images_paths = fold[1]
    suspect_image = open_suspect(suspect_image_path)
    control_images = open_suspect(control_images_paths)

    suspect_prnu = prnu(suspect_image)
    control_prnu = prnu(control_images)

    cross_cor = crosscorr_2d(suspect_prnu, control_prnu)
    pce_val = pce(cross_cor)
    print(pce_val['cc'])

def main(mode):
    correct = 0
    incorrect = 0
    excluded = ''
    tp = 0
    fp = 0
    tn = 0
    fn = 0
    
    if mode == 'No Match':
        load_equal_data(50)
        folds, excluded = create_no_match_fold()
        get_fold_stats(folds)
        folds = choose_no_match_suspect(folds, excluded)
    elif mode == 'Equal Size':
        load_equal_data(50)
        folds = create_equal_sized_folds()
        get_fold_stats(folds)
        folds = choose_suspect_image(folds) # folds[0] will contain ['suspect image path', [list of control image paths]]
    elif mode == 'Fixed Size':
        load_data()
        folds = create_fixed_size_fold(300)
        get_fold_stats(folds)
        folds = choose_suspect_image(folds) # folds[0] will contain ['suspect image path', [list of control image paths]]
    

    print("-----------------------------------------------------------------------------------------------------------")
    for fold in folds:
        to_write = [] # path (of suspect image), actual class, predicted class, class scores
        control_fingerprints = []
        scores = [[],[],[],[],[],[]]
        pce_store = [[],[],[],[],[],[]]
        avg_scores = []
        suspect_image_path = fold[0]
        control_images_paths = fold[1]
        suspect_image = open_suspect(suspect_image_path)
        control_images = open_control(control_images_paths)


        suspect_prnu = prnu(suspect_image)
        prog = len(control_images)
        counter = 1
        for image in control_images:
            image = np.asarray(image)
            image = cut_ctr(image, (image.shape))
            control_fingerprints.append(prnu(image))
            #print(f"Progress: {round((counter/prog) * 100, 2)}%")
            counter+=1

        suspect_fingerprint = suspect_prnu
        count = 0
        tracker = 1
        for fingerprint in control_fingerprints:
            cross_cor = crosscorr_2d(suspect_fingerprint, fingerprint)
            pce_val = pce(cross_cor)
            pos = devices[labels[control_images_paths[count]]]
            scores[pos].append(pce_val['cc'])
            pce_store[pos].append(pce_val['pce'])
            count += 1
            #print(f" Cross Correlation Progress: {round((tracker/prog) * 100, 2)}%")
            tracker+=1
        
        for i in range(6):
            avg_scores.append(average(scores[i]))

        flat_pce = [value for sublist in pce_store for value in sublist]
        max_pce = max(flat_pce)
        if max_pce > 60:
            index_max = avg_scores.index(max(avg_scores))
            predicted_identity = (devices[index_max])
        else:
            predicted_identity = 'Not In Set'

        if mode == 'No Match':
             print(f"Suspect Device : {labels[suspect_image_path]} (Not In Set)  \nProgram Predicted Match: {predicted_identity}")
        else:
            print(f"Suspect Device : {labels[suspect_image_path]}  \nProgram Predicted Match: {predicted_identity}")
        print(f"Scores: \nD3500: {average(scores[0])}   Drone: {average(scores[1])}   iPhone 8: {average(scores[2])}   iPhone 13: {average(scores[3])}   Samsung Galaxy S7 A: {average(scores[4])}   Samsung Galaxy S7 B: {average(scores[5])}")
        print("--------------------------------------------------------------------------------------------------------------------------------------------------------------")
        if max_pce > 60:
            if labels[suspect_image_path] == predicted_identity:
                tp += 1
                verdict = "TP"
            else:
                fp += 1
                verdict = "FP"
        else:
            if predicted_identity  == 'Not In Set' and labels[suspect_image_path] == excluded:
                tn += 1
                verdict = "TN"
            else:
                fn += 1
                verdict = "FN"
        to_write.append([suspect_image_path, labels[suspect_image_path], predicted_identity, average(scores[0]), average(scores[1]), average(scores[2]), average(scores[3]), average(scores[4]), average(scores[5]), verdict, mode])
        full_dataset.append(to_write[0])
    return tp, tn, fp, fn

def test_no_match():
    correct = 0
    incorrect = 0
    tp = 0
    fp = 0
    tn = 0
    fn = 0

    for i in range(25):
        # Testing mode 1
        start_time = time.time()
        scores = main('No Match') # tp, tn, fp, fn
        correct += scores[0]
        tp += scores[0]
        correct += scores[1]
        tn += scores[1]
        incorrect += scores[2]
        fp += scores[2]
        incorrect += scores[3]
        fn += scores[3]
        batch_correct = scores[0] + scores[1]
        batch_incorrect = scores[2] + scores[3]
        succesrate = (batch_correct/(batch_correct+batch_incorrect))*100
        print(f"Success Rate: {succesrate}%")
        end_time = time.time()
        elapsed = end_time - start_time
        full_dataset.append(['Batch Accuracy :', succesrate, '-', '-', '-', '-', '-', '-', 'Execution Time (Minutes)',elapsed/60])

    final_accuracy = (correct/(correct+incorrect))*100
    full_dataset.append(['Test Round Accuracy :', final_accuracy, '-', '-', '-', '-', f'TP: {tp}', f'FP: {fp}', f'TN: {tn}',f'FN: {fn}',])

    workbook = pyxl.load_workbook('new_data.xlsx')
    sheet = workbook['No Match']
    workbook.active = sheet

    for row_index, row_data in enumerate(full_dataset, start=1):
        for column_index, value in enumerate(row_data, start=1):
            sheet.cell(row=row_index, column=column_index).value = value

    workbook.save('new_data.xlsx')
    full_dataset.clear()
    full_dataset.append(['Image Path', 'Actual Class', 'Predicted Class', 'D3500', 'Drone', 'iPhone 8', 'iPhone 13', 'Galaxy S7 A', 'Galaxy S7 B', 'Result', 'Mode'])
    
def test_equal_size():
    correct = 0
    incorrect = 0
    tp = 0
    fp = 0
    tn = 0
    fn = 0

    for i in range(25):
        # Testing mode 2
        start_time = time.time()
        scores = main('Equal Size') # tp, tn, fp, fn
        correct += scores[0]
        tp += scores[0]
        correct += scores[1]
        tn += scores[1]
        incorrect += scores[2]
        fp += scores[2]
        incorrect += scores[3]
        fn += scores[3]
        batch_correct = scores[0] + scores[1]
        batch_incorrect = scores[2] + scores[3]
        succesrate = (batch_correct/(batch_correct+batch_incorrect))*100
        print(f"Success Rate: {succesrate}%")
        end_time = time.time()
        elapsed = end_time - start_time
        full_dataset.append(['Batch Accuracy :', succesrate, '-', '-', '-', '-', '-', '-', 'Execution Time (Minutes)',elapsed/60])

    final_accuracy = (correct/(correct+incorrect))*100
    full_dataset.append(['Test Round Accuracy :', final_accuracy, '-', '-', '-', '-', f'TP: {tp}', f'FP: {fp}', f'TN: {tn}',f'FN: {fn}',])

    workbook = pyxl.load_workbook('new_data.xlsx')
    sheet = workbook['Equal Size']
    workbook.active = sheet

    for row_index, row_data in enumerate(full_dataset, start=1):
        for column_index, value in enumerate(row_data, start=1):
            sheet.cell(row=row_index, column=column_index).value = value

    workbook.save('new_data.xlsx')
    full_dataset.clear()
    full_dataset.append(['Image Path', 'Actual Class', 'Predicted Class', 'D3500', 'Drone', 'iPhone 8', 'iPhone 13', 'Galaxy S7 A', 'Galaxy S7 B', 'Result', 'Mode'])

def test_fixed():
    correct = 0
    incorrect = 0
    tp = 0
    fp = 0
    tn = 0
    fn = 0

    for i in range(25):
        # Testing mode 3
        start_time = time.time()
        scores = main('Fixed Size') # tp, tn, fp, fn
        correct += scores[0]
        tp += scores[0]
        correct += scores[1]
        tn += scores[1]
        incorrect += scores[2]
        fp += scores[2]
        incorrect += scores[3]
        fn += scores[3]
        batch_correct = scores[0] + scores[1]
        batch_incorrect = scores[2] + scores[3]
        succesrate = (batch_correct/(batch_correct+batch_incorrect))*100
        print(f"Success Rate: {succesrate}%")
        end_time = time.time()
        elapsed = end_time - start_time
        full_dataset.append(['Batch Accuracy :', succesrate, '-', '-', '-', '-', '-', '-', 'Execution Time (Minutes)',elapsed/60])

    final_accuracy = (correct/(correct+incorrect))*100
    full_dataset.append(['Test Round Accuracy :', final_accuracy, '-', '-', '-', '-', f'TP: {tp}', f'FP: {fp}', f'TN: {tn}',f'FN: {fn}',])

    workbook = pyxl.load_workbook('new_data.xlsx')
    sheet = workbook['Fixed Size']
    workbook.active = sheet

    for row_index, row_data in enumerate(full_dataset, start=1):
        for column_index, value in enumerate(row_data, start=1):
            sheet.cell(row=row_index, column=column_index).value = value

    workbook.save('new_data.xlsx')
    full_dataset.clear()
    full_dataset.append(['Image Path', 'Actual Class', 'Predicted Class', 'D3500', 'Drone', 'iPhone 8', 'iPhone 13', 'Galaxy S7 A', 'Galaxy S7 B', 'Result', 'Mode'])

test_started = time.time()
now = datetime.now()
current_time = now.strftime("%H:%M:%S")
print(f"Test Started At: {current_time}")
#test_no_match()
test_equal_size()
#test_fixed()
test_finished = time.time()
now = datetime.now()
current_time = now.strftime("%H:%M:%S")
print(f"Test Finished At: {current_time}")
print(f"Test Completed In {(test_finished-test_started)/60} Minutes")